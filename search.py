import os
import re
import sys
import docx
from openpyxl import load_workbook
from pypdf import PdfReader

class FileScanner:
    def __init__(self, path):
        self.path = path
        self.search_term = ""
        self.files = []
        self.match_count = 0

    def set_path(self, path):
        """Set the directory path."""
        self.path = path
        if path:
            os.chdir(path)
        self.files = list(enumerate(os.listdir(), 1))

    def set_search_term(self, search_term):
        """Set the search term."""
        self.search_term = search_term

    def scan_files(self):
        """Scan all files in the directory."""
        self.match_count = 0
        for index, file in self.files:
            self.display_progress(file, index)
            self.scan_file(file, index)
        
        print(f"\nNumber of matches found: {self.match_count} in {len(self.files)} files.")

    def scan_file(self, file, index):
        """Scan a single file for the search term."""
        try:
            if file.endswith(".txt"):
                self.scan_txt(file)
            elif file.endswith(".docx"):
                self.scan_docx(file)
            elif file.endswith(".pdf"):
                self.scan_pdf(file)
            elif file.endswith(".xlsx"):
                self.scan_xlsx(file)
        except Exception as e:
            print(f"\nError scanning {file}: {e}")

    def scan_txt(self, file):
        """Scan a .txt file."""
        with open(file, 'r') as f:
            for line_number, line in enumerate(f, start=1):
                self.search_in_content(file, line, line_number)

    def scan_docx(self, file):
        """Scan a .docx file."""
        doc = docx.Document(file)
        for line_number, paragraph in enumerate(doc.paragraphs, start=1):
            self.search_in_content(file, paragraph.text, line_number)

    def scan_pdf(self, file):
        """Scan a .pdf file."""
        pdf = PdfReader(file)
        for page_number, page in enumerate(pdf.pages, start=1):
            self.search_in_content(file, page.extract_text(), page_number, is_page=True)

    def scan_xlsx(self, file):
        """Scan an .xlsx file."""
        workbook = load_workbook(file)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row, col)
                    if cell.value and self.search_term in str(cell.value):
                        print(f"\nFound in {file} in sheet '{sheet.title}', cell {cell.coordinate}")

    def search_in_content(self, file, content, position, is_page=False):
        """Search for the term in given content."""
        matches = re.finditer(self.search_term, content)
        for match in matches:
            self.match_count += 1
            if is_page:
                print(f"\nFound in {file} at page {position}, character {match.start() + 1}")
            else:
                print(f"\nFound in {file} at line {position}, character {match.start() + 1}")

    def display_progress(self, file, index):
        """Display progress of scanning."""
        sys.stdout.write(f"\rScanning: {file} ({index}/{len(self.files)})")
        sys.stdout.flush()


# Main Execution
def main():
    scanner = FileScanner("")
    while True:
        if input("\nEnter 1 to change location or any key to continue: ") == '1':
            path = input("Enter folder location: ")
            scanner.set_path(path)
        
        search_term = input("Enter the word to search: ")
        scanner.set_search_term(search_term)

        scanner.scan_files()

        if input("\nEnter 1 to scan again or any key to exit: ") != '1':
            break


if __name__ == "__main__":
    main()
